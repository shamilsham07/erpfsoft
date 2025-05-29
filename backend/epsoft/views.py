from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny 
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import UserAccount
from .models import Staff
from.models import staffwage
from django.db.models import Sum
from.models import epfcalculation
import math
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl import Workbook
from .serializer import MyModelSerializer
from datetime import datetime, timedelta
from django.core.exceptions import ObjectDoesNotExist
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth.hashers import make_password, check_password

# Create your views here.



@api_view(["POST"])
@permission_classes([AllowAny]) 
def userlogin(request):
    print("heelo")
    data=request.data
    username=data.get("username")
    password=data.get("password")
    print(username)
    print(password)

    try:
        user=UserAccount.objects.get(username=username)
        print(user.password)
        hashedpass=check_password(password,user.password)
        print(hashedpass)
        if(hashedpass):
         return JsonResponse ({"message":"recieved successfull"})
        else:
            return JsonResponse({"no-user":"no-userFound"})   
    except ObjectDoesNotExist:
        print("jabar")
        return JsonResponse({"no-user":"no-userFound"})
        

@api_view(["POST"])
@permission_classes([AllowAny]) 
def createaccount(request):
    print("jiiii")
    try:
        data=request.data
        username=data.get("username")
        password=data.get("password")
        email=data.get("email")
        phonenumber=data.get("phonenumber")
        hashedpass=make_password(password)
        print(hashedpass)
        if UserAccount.objects.filter(email=email).exists():
            return JsonResponse({"email":'alreadyexists'})
        if UserAccount.objects.filter(phone=phonenumber).exists():
            return JsonResponse({"phonenumber":"alreadyexists"})
        UserAccount.objects.create(username=username,is_admin=True,email=email,phone=phonenumber,password=hashedpass)
        return JsonResponse({"message":"usercraeted"})
    except Exception as e:
        print(e)
        return JsonResponse({"wrong":"error"})
    
    
@csrf_exempt
@api_view(["POST"])
@authentication_classes([])  # Disable authentication
@permission_classes([AllowAny])
def savefile(request):
    datas = request.data.get("data")
    try:
       for items in datas:        
            Staff.objects.create(
                Name=items.get("Name (Alphabets Only)"),
                
                UAN=items.get("UAN"),
                dob=items.get("Date of Birth (DD/MM/YYYY) Eg : 01/01/1970"),
                DateOfAppointment=items.get("Date of Appointment (DD/MM/YYYY)Eg : 01/01/1970"),    
                aadhar=items.get("Aadhar"),
                inputnumber=items.get("IP Number (10 Digits Only)"),
                gender=items.get("Gender (M/F/TG)"),
                
            )
       return JsonResponse({"message": "good"})
    except Exception as e:
        print(e)
        return JsonResponse({"error": str(e)})
    
    
@api_view(["GET"])
@authentication_classes([])  # Disable authentication
@permission_classes([AllowAny])
def getdata(request):
    try:
        search=request.GET.get("search",'')
        staff=Staff.objects.all().order_by("id")
        staffcount=staff.count()   
        
        if search:
            print("ff",search)
            staff=staff.filter(Name__icontains=search)or None
            print("value",staff)
        else:
            staff=Staff.objects.all().order_by("id")
          
        print(staff)
        if staff:
            paginator=PageNumberPagination()
            paginator.page_size=10
            page=paginator.paginate_queryset(staff,request)
            if page is not None:
                serializer=MyModelSerializer(page,many=True)
                print("jabar")
                pageSize=paginator.page_size
                print(serializer.data)
                totalpages=math.ceil(staffcount/pageSize)
                
                page_number = request.GET.get('page', 1)   
                print(page_number)
                return paginator.get_paginated_response({"data":serializer.data,"number":page_number,"totalpages":totalpages})
            serializer=MyModelSerializer(staff,many=True)
            print("hussain")
            print(serializer.data)
            return JsonResponse({"data":serializer.data},safe=False)
                
            
        else:
            return JsonResponse({"error":'wrong'})
    except Exception as e:
        print(e)
        return JsonResponse({"wrong":"very bad"})
   
   
@api_view(["POST"])
@authentication_classes([])  # Disable authentication
@permission_classes([AllowAny])
def addstaff(request):
    try:
        data=request.data
        name=data.get("name",None)
        aadhar=data.get("aadhar",None)
        ifsc=data.get("ifsccode",None)
        accno=data.get("accno",None)
        print(accno)
        print(ifsc)
        ipno=data.get("ipno",None)
        gender=data.get("gender",None)
        uan=data.get("uan",None)
        doa=data.get("doa",None)
        dob=data.get("dob",None) 
        Staff.objects.create(
        Name=name,
        dob=dob,
        DateOfAppointment=doa,
        inputnumber=ipno,
        gender=gender,
        UAN=uan,
       
        aadhar=aadhar,     
        )      
        return JsonResponse({"message":"good"})
    except Exception as e:
        return JsonResponse({"worng":"something went wrong"})
    
@api_view(["POST"])    
@authentication_classes([])      
@permission_classes([AllowAny])
def getthedataforupdate(request):
    try:
        print("hello")
        data=request.data
        id=data.get("id")
        print(id)
        print("........................................")
        staff=Staff.objects.get(id=id)
        staffdata={
            "Name":staff.Name,
            "dob":staff.dob,
            "UAN":staff.UAN,
            "aadhar":staff.aadhar,
            "gender":staff.gender,
            "doa":staff.DateOfAppointment,
            "inputnumber":staff.inputnumber,
            "accountnumber":staff.accountnumber,
            "ifsccode":staff.ifsccode,
             "id":id,
               
              
        }
        return JsonResponse({"data":staffdata},status=200)
    except Exception as e:
        return JsonResponse({"bad":"agly"})         
     
  
  
    
@api_view(["POST"])    
@authentication_classes([])      
@permission_classes([AllowAny])
def updtaingthestaff(request):
    try:
        data=request.data
        print(data)
        id=data.get("id")
        print(id)
        name=data.get("name")
        ipno=data.get("ipno")
        uan=data.get("UAN")
        aadhar=data.get('aadhar')
        gender=data.get("gender")
        accno=data.get("accno")
        ifsccode=data.get("ifsccode")
        print(name)
        
        if id:
            print("hellll")
            staff=Staff.objects.get(id=id)
            staff.Name=name
            staff.UAN=uan
            staff.inputnumber=ipno
            staff.accountnumber=accno
            staff.gender=gender
            staff.ifsccode=ifsccode
            staff.aadhar=aadhar
            staff.save()
            
        
        return JsonResponse({"datta":"good"},status=200)
    except Exception as e:
        print(e)
        return JsonResponse({"error":"good"})
    
@api_view(["POST"])    
@authentication_classes([])      
@permission_classes([AllowAny])
def updateBank(request):
    try:
        data=request.data.get("data")     
        print("jaber")
        print("hello",data)
        

        for items in data:
            name=items.get("NAME")
            print(name)   
            accno=items.get("accnumber")
            print(accno)
            ifsc=items.get("IFS CODE")
            if Staff.objects.filter(Name=name).exists():
                print("")
                staff = Staff.objects.get(Name=name)
                print(staff)
                print(staff.UAN)
                print(staff)
                print(staff.accountnumber)
                print(staff.ifsccode)
                staff.accountnumber=accno
                staff.ifsccode=ifsc
                staff.save()      
                

        return JsonResponse({"message":"good"})
    except Exception as e:
        print(e)
        return JsonResponse({"bad":"badsss"})
    
    
@api_view(["POST"])
@authentication_classes([])  # Disable authentication
@permission_classes([AllowAny])
def deletethestaff(request):
    try:
        data=request.data
        deleteid=data.get("deleteid")
        print("//",deleteid)
        if deleteid:
            Staff.objects.delete(id=deleteid)
            staff=Staff.objects.all()
            serializer=MyModelSerializer(staff,many=True)
            
            return JsonResponse({"data":serializer.data},safe=False)
    except Exception as e:
        print(e)
        return JsonResponse({"error":"something fishy"})
         
        

@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])
def gettotaldetails(request):
    try:
        data=request.data
        date=data.get("date")
        print(date)
        wageData = []
        if staffwage.objects.filter(date=date).exists():
            staffs=staffwage.objects.filter(date=date).order_by("id")
            print(staffs)    
            print("hi")
            for staff in staffs:
                staffName=Staff.objects.get(id=staff.staffid_id)
                
                staff_wage = epfcalculation.objects.get(wage=staff)
              
                wageData.append({
                    'id':staff.id,
                    "gross":staff.wage,
                    'epf':staff_wage.epf,
                    "er":staff_wage.er,
                    "eps":staff_wage.eps,
                    "edli":staff_wage.edli,
                    "ee":staff_wage.ee,
                    'name':staffName.Name,
                    'uan':staffName.UAN,
                    'eps_employer':staff_wage.eps_employer,
                    "attendence":staff.staffattendence,
                    "ncp_days":staff_wage.ncp_days
                    
                    
                })
               
            return JsonResponse({"data":wageData})
        
        else:
            print("shamil")
            print(date)
            staff=Staff.objects.all().order_by("id")    
            print("tghe",staff)
           
            wagedata=[]
            for staffs in staff:
                wagedata.append({
                    "name":staffs.Name,
                    "uan":staffs.UAN,
                    "gross":0,
                    "er":0,
                    "epf":0,
                    "eps":0,
                    "edli":0,
                    "eps":0,
                    'eps_employer':0,
                    "id":staffs.id,
                    "attendence":0,
                    "ncp_days":0

                })
            return JsonResponse({"afterdate":wagedata},)
    except Exception as e:
        print(e)
        return JsonResponse({"error":"something fishy"})     
    

@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])  
def SaveTheWage(request):
    try:
        print("suluikka")
        data=request.data
        date=data.get("date")
        print("data",data)
        print(date)
        return JsonResponse({"message":"good"})
    except Exception as e:
        return JsonResponse({"error":"wrong"})

@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])  
def savingorupdatewage(request):
    try:
        data=request.data
        print("hellooo")
        datas=data.get("details")
        print(datas)
        getdate=data.get("date")
        print("    ",getdate)
        date=staffwage.objects.filter(date=getdate).exists()
        print("shamil",date)
        if date:    
            for items in datas:
                id=items.get("id")
                gross=items.get("gross")or 0
                epf=items.get("epf")or 0
                er=items.get("er")or 0
                eps=items.get("eps")or 0
                edli=items.get("edli")or 0
                ncp_days=items.get("ncp_days")
                attendence=items.get("attendence")
                ee=items.get("ee")or 0
                eps_employer=items.get("eps_employer")or 0
                
                
                if id:
                    if staffwage.objects.filter(id=id).exists():
                        staff=staffwage.objects.get(id=id)
                        staff.staffattendence=attendence
                        staff.wage=gross
                        staff.save()
                        staff_wage=epfcalculation.objects.get(wage=staff)
                        staff_wage.epf=epf
                        staff_wage.eps=eps
                        staff_wage.er=er
                        staff_wage.edli=edli
                        staff_wage.ee=ee
                        staff_wage.ncp_days=ncp_days
                        staff_wage.eps_employer=eps_employer
                        staff_wage.save()
                
            return JsonResponse({"good":"baddddddddddddddddddddd"})   
        
               
               
        else:
            for items in datas:
                id=items.get("id") or 0
                epf=items.get("epf")or 0
                eps=items.get("eps")or 0
                edli=items.get("edli")or 0
                ee=items.get("ee")or 0
                eps_employer=items.get("eps_employer")or 0
                er=items.get("er")or 0
                ncp_days=items.get("ncp_days")or 0
                attendence=items.get("attendence")or 0


                
                
                name=items.get("Name",None) 
                wage = items.get("gross",0)
                if wage:
                    wage = int(wage) 
                else:
                    wage = 0
                if id:
                    staffwage.objects.create(staffid_id=id,wage=wage,date=getdate,staffattendence=attendence)
                    wages=staffwage.objects.get(staffid_id=id,date=getdate)
                    theid=wages.id
                    print("ererer",theid)
                    epfcalculation.objects.create(staffid_id=id,wage_id=theid,date=getdate,epf=epf,eps=eps,edli=edli,ee=ee,eps_employer=eps_employer,er=er,ncp_days=ncp_days)
                    
                    
        return  JsonResponse({"data":"mesage recieved succesfuuly"})
    except Exception as e:
        print(e)
        return JsonResponse({"eroor":"errrororororo"})
    
    
    
    
@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])  
def goandfetchdata(request):
    try:
        print("hello")
        data=request.data
        date=data.get("date")
        print(date)
        staff=staffwage.objects.filter(date=date).order_by("id")or None
        print(staff)
        wagedata=[]
        if staff:
            for staffs in staff:
                id=staffs.staffid_id
                print(id)
                if id:
                    staff_detail=Staff.objects.get(id=id)
                    name=staff_detail.Name  
                    uan=staff_detail.UAN 
                wage=staffs.wage
                staff_wage=epfcalculation.objects.get(wage=staffs)  
                print("helll")     
                print(staff_wage)
                epf=staff_wage.epf
                print(epf)
                eps=staff_wage.eps
                edli=staff_wage.edli   
                ee=staff_wage.ee
                eps_employer=staff_wage.eps_employer   
                er=staff_wage.er
                
                ncp_days=staff_wage.ncp_days
                refund=staff_wage.refund
                
                wagedata.append({
                    "name":name,
                    "uan":uan,
                    "date":date,
                    "gross":wage,
                    "epf":epf,
                    "eps":eps,
                    "edli":edli,
                    "ee":ee,
                    "eps_employer":eps_employer,
                    "er":er,
                    "ncp_days":ncp_days,
                    "refund":refund,
                    "id":staff_wage.id
                    
                    
                })
            
            print("wagedate",wagedata)        
            return JsonResponse({"data":wagedata})    
                
                
            
        else:
            print("helloo")
        
        return JsonResponse({"message":"great"})
    except Exception as e:
        print(e)
        return JsonResponse({"error":"something went wrong"})
    
 
@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])     
def calenderdata(request):
    try:
        data=request.data
        date=data.get("date")
        print(date)
        
        if staffwage.objects.filter(date=date).exists():
            staffs=staffwage.objects.filter(date=date).order_by("id")      
            wagedata=[]
            index=0
            for staff in staffs:
                gross=staff.wage
                date=staff.date
                staff_wage=epfcalculation.objects.get(wage=staff)
                print(staff_wage)
                print(staff_wage.epf)
                epf=staff_wage.epf
                eps=staff_wage.eps
                edli=staff_wage.edli   
                ee=staff_wage.ee
                eps_employer=staff_wage.eps_employer   
                er=staff_wage.er
                ncp_days=staff_wage.ncp_days
                refund=staff_wage.refund
                index=index+1
                
                staff_obj=staff.staffid
                
                wagedata.append({
                    "name":staff_obj.Name,     
                    "epf":epf,
                    "eps":eps,
                    "gross":gross,
                    "date":date,
                    "edli":edli,
                    "ee":ee,
                    "eps_employer":eps_employer,
                    "er":er,
                    "ncp_days":ncp_days,
                    "refund":refund,
                    "index":index
                    
                })
                
            return JsonResponse({"data":wagedata})
            
        else:
            print("hhhh")
            return JsonResponse({"error":"something went wrong"})
    except Exception as e:
        print(e)
        return JsonResponse({"error":"something went wrong"})
    
@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])     
def onDownload(request):
    data=request.data
    date=data.get("date")
    sheet=data.get("sheet")
    print(sheet)
    print(date)
    if staffwage.objects.filter(date=date).exists():
      if sheet=="sheet1":  
        headers = [
                "UAN", "MEMBERS NAME", "GROSS WAGES", "EPF WAGES", "EDLI WAGES",
                "EPF CONTRIBUTED", "EPS CONTRIBUTED", "ER", "NCP_DAYS", "REFUND"
            ]
        wb = Workbook()
        ws = wb.active
        print("hello")
        staff_details=staffwage.objects.filter(date=date).order_by('id')
        print(staff_details)
        details=[]
        for staff in staff_details:
            
            gross=staff.wage
            name_details=staff.staffid
            name=name_details.Name   
            wage_details=epfcalculation.objects.get(wage=staff)
            print(gross)   
            
            print(name)
            details.append([name_details.UAN,name,gross,wage_details.epf,wage_details.edli,wage_details.ee,wage_details.er,wage_details.ncp_days,wage_details.refund])
            
     
         
        
      else:    
            print("hhhhh") 
            wb = Workbook()
            ws = wb.active 
            staff_details=staffwage.objects.filter(date=date).order_by("id")
            headers=[
                    "input number","Name","attendence","wage","reason"
                ]
            details=[]
            for staff in staff_details:
                print(staff.staffid)
                name=staff.staffid.Name
                ip=staff.staffid.inputnumber
                attendence=staff.staffattendence
                gross=staff.wage
                if attendence:
                    reason=""
                else:
                    reason=1
                details.append([
                    ip,name,attendence,gross,reason
                ])
            
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = max(len(column_title) + 2, 15)

    # Write data rows
    for row in details:
        ws.append(row)

    ws.freeze_panes = "A2"

    # Save workbook
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(
        file_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample.xlsx"'
    return response
     
    
          
          
@api_view(["POST"])
@authentication_classes([]) 
@permission_classes([AllowAny])     
def getcountofusers(request):
    data=request.data
    date=data.get("date")
    print(date)
    staff=[]
    
    try:
        if date:
                staff_details=staffwage.objects.filter(date=date)
                print(staff_details)
               
                total_epf = 0
                total_attendence=0
                for staffs in staff_details:
                    total_attendence+=int(staffs.staffattendence)
                    wage_details=epfcalculation.objects.get(wage=staffs)
                   
                    total_epf += wage_details.epf
               
                staff.append({  
                "wage": staffwage.objects.filter(date=date).aggregate(total=Sum("wage"))['total'],
                "total_epf":total_epf,
                "attendence":total_attendence,
               
            
                })
                print(staff)
                
        data=Staff.objects.all().count()
        print(data)
        return JsonResponse({"message":data,"totalwage":staff})
    except Exception as e:
        print(e)
        return JsonResponse({"error":"something went wrong"})
  
     
    